"""Unit tests for the create-spreadsheet financial-model template.

The template lives in `app/agents/skills/builtin/create-spreadsheet/templates/
report.py` — a sandbox-executed asset, not an importable API module (no
``__init__.py``, hyphenated directory). It is loaded here the same way the
sandbox runs it: straight off the file system, via importlib. openpyxl is a
dev-group dependency because the module (and its tests) need it; the API
runtime never imports it.
"""

import importlib.util
from pathlib import Path
import sys
from unittest.mock import patch

from openpyxl import Workbook, load_workbook
from openpyxl.chart._chart import ChartBase
from openpyxl.chart.title import Title
from openpyxl.formatting.rule import Rule
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet
import pytest

# The file sits 5 levels below the repo root in the live tree, but the
# mutation lane's workdir adds one more level — walk up to the tree marker
# (apps/api/app) so the path resolves in both layouts.
REPO_ROOT = next(
    (p for p in Path(__file__).resolve().parents if (p / "apps" / "api" / "app").is_dir()),
    Path(__file__).resolve().parents[5],
)
_REPORT_PATH = (
    REPO_ROOT
    / "apps"
    / "api"
    / "app"
    / "agents"
    / "skills"
    / "builtin"
    / "create-spreadsheet"
    / "templates"
    / "report.py"
)
_SPEC = importlib.util.spec_from_file_location("report_template", _REPORT_PATH)
assert _SPEC is not None and _SPEC.loader is not None
report = importlib.util.module_from_spec(_SPEC)
_SPEC.loader.exec_module(report)

ASSUMPTIONS = report.ASSUMPTIONS


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _rgb(fill: PatternFill) -> str:
    """Normalize an openpyxl ARGB (e.g. "00F2F6FB") to the bare 6-hex code."""
    rgb = fill.fgColor.rgb or ""
    return rgb[2:] if len(rgb) == 8 and rgb.startswith("00") else rgb


def _fill(ws: Worksheet, ref: str) -> str:
    return _rgb(ws[ref].fill)


def _merged(ws: Worksheet) -> set[str]:
    return {str(r) for r in ws.merged_cells.ranges}


def _rules_by_range(ws: Worksheet) -> dict[str, list[Rule]]:
    return {str(cf.sqref): cf.rules for cf in ws.conditional_formatting}


def _rule(ws: Worksheet, range_: str) -> Rule:
    rules = _rules_by_range(ws)[range_]
    assert len(rules) == 1
    return rules[0]


def _title_text(title: Title) -> str:
    return title.tx.rich.p[0].r[0].t


def _series_range(chart: ChartBase, index: int) -> str:
    return chart.series[index].val.numRef.f


def _row_by_label(model: dict, label: str) -> tuple:
    for row in model["rows"]:
        if row[0] == label:
            return row
    raise AssertionError(f"no row labeled {label!r} in model")


def _margin_by_label(model: dict, label: str) -> tuple:
    for row in model["margin_rows"]:
        if row[0] == label:
            return row
    raise AssertionError(f"no margin row labeled {label!r} in model")


def _patched_assumptions(**overrides: object) -> dict:
    a = dict(ASSUMPTIONS)
    a.update(overrides)
    return a


@pytest.fixture
def ws() -> Worksheet:
    return Workbook().active


@pytest.fixture
def model() -> dict:
    return report.compute_model()


# ---------------------------------------------------------------------------
# compute_model — the pure P&L math
# ---------------------------------------------------------------------------


class TestComputeModel:
    def test_years_are_consecutive_starting_at_start_year(self, model: dict) -> None:
        assert model["years"] == [2025, 2026, 2027, 2028, 2029]

    def test_revenue_compounds_at_growth_rate(self, model: dict) -> None:
        revenue = model["revenue"]
        expected = [4_800_000.0 * 1.28**i for i in range(5)]
        assert revenue == pytest.approx(expected)

    def test_cogs_is_fixed_pct_of_revenue(self, model: dict) -> None:
        _, cogs, _, _ = _row_by_label(model, "COGS")
        assert cogs == pytest.approx([r * 0.42 for r in model["revenue"]])

    def test_gross_profit_is_revenue_minus_cogs(self, model: dict) -> None:
        _, gp, _, _ = _row_by_label(model, "Gross Profit")
        assert gp == pytest.approx(
            [r - c for r, c in zip(model["revenue"], _row_by_label(model, "COGS")[1])]
        )

    def test_headcount_rounds_compounding_headcount(self, model: dict) -> None:
        assert model["headcount"] == [32, 38, 45, 53, 62]

    def test_personnel_is_headcount_times_cost_per_head(self, model: dict) -> None:
        _, personnel, _, _ = _row_by_label(model, "  Personnel")
        assert personnel == pytest.approx(
            [32 * 145_000.0, 38 * 145_000.0, 45 * 145_000.0, 53 * 145_000.0, 62 * 145_000.0]
        )

    def test_opex_lines_are_pct_of_revenue(self, model: dict) -> None:
        _, sm, _, _ = _row_by_label(model, "  Sales & Marketing")
        _, rd, _, _ = _row_by_label(model, "  Research & Development")
        _, ga, _, _ = _row_by_label(model, "  General & Admin")
        revenue = model["revenue"]
        assert sm == pytest.approx([r * 0.16 for r in revenue])
        assert rd == pytest.approx([r * 0.11 for r in revenue])
        assert ga == pytest.approx([r * 0.07 for r in revenue])

    def test_total_opex_sums_every_line_per_year(self, model: dict) -> None:
        _, total, _, _ = _row_by_label(model, "Total Opex")
        expected = [
            6_272_000.0,
            7_598_960.0,
            9_198_868.8,
            11_107_552.064,
            13_370_866.641_92,
        ]
        assert total == pytest.approx(expected)

    def test_ebitda_ebit_tax_net_income_exact_values(self, model: dict) -> None:
        _, ebitda, _, _ = _row_by_label(model, "EBITDA")
        _, ebit, _, _ = _row_by_label(model, "EBIT")
        _, tax, _, _ = _row_by_label(model, "Tax")
        _, net, _, _ = _row_by_label(model, "Net Income")
        depreciation = [240_000.0] * 5

        assert ebitda == pytest.approx(
            [-3_488_000.0, -4_035_440.0, -4_637_563.2, -5_269_080.896, -5_897_623.546_88]
        )
        assert ebit == pytest.approx([e - d for e, d in zip(ebitda, depreciation)])
        assert tax == [0.0] * 5
        assert net == pytest.approx([e - t for e, t in zip(ebit, tax)])
        assert model["net_income"] == pytest.approx(net)

    def test_depreciation_is_flat_each_year(self, model: dict) -> None:
        _, dep, _, _ = _row_by_label(model, "Depreciation & Amort.")
        assert dep == [240_000.0] * 5

    def test_rows_are_in_pnl_order_with_indented_opex(self, model: dict) -> None:
        labels = [row[0] for row in model["rows"]]
        assert labels == [
            "Revenue",
            "COGS",
            "Gross Profit",
            "  Personnel",
            "  Sales & Marketing",
            "  Research & Development",
            "  General & Admin",
            "Total Opex",
            "EBITDA",
            "Depreciation & Amort.",
            "EBIT",
            "Tax",
            "Net Income",
        ]
        for row in model["rows"]:
            assert len(row[1]) == 5
            assert row[2] == report.FMT_CURRENCY

    def test_subtotal_rows_are_bold(self, model: dict) -> None:
        for label in ("Revenue", "Gross Profit", "Total Opex", "EBITDA", "EBIT", "Net Income"):
            assert _row_by_label(model, label)[3] is True
        for label in ("COGS", "Tax", "Depreciation & Amort.", "  Personnel"):
            assert _row_by_label(model, label)[3] is False

    def test_margins_are_ratios_computed_in_python(self, model: dict) -> None:
        _, gross = _margin_by_label(model, "Gross Margin")
        _, ebitda = _margin_by_label(model, "EBITDA Margin")
        _, net = _margin_by_label(model, "Net Margin")
        assert gross == pytest.approx([0.58] * 5)
        revenue = model["revenue"]
        _, gross_profit, _, _ = _row_by_label(model, "Gross Profit")
        _, total_opex, _, _ = _row_by_label(model, "Total Opex")
        assert ebitda == pytest.approx(
            [(gp - o) / r for gp, o, r in zip(gross_profit, total_opex, revenue)]
        )
        assert net == pytest.approx([ni / r for ni, r in zip(model["net_income"], revenue)])

    def test_summary_exact_values(self, model: dict) -> None:
        summary = model["summary"]
        assert summary["Total Revenue (5y)"] == pytest.approx((41_759_551.488, report.FMT_CURRENCY))
        assert summary["Total Net Income (5y)"] == pytest.approx(
            (-24_527_707.642_88, report.FMT_CURRENCY)
        )
        assert summary["Revenue CAGR"] == pytest.approx((0.28, report.FMT_PERCENT))
        assert summary["Avg Gross Margin"] == pytest.approx((0.58, report.FMT_PERCENT))
        assert summary["Avg Net Margin"] == pytest.approx(
            (-0.623_274_718_523_025_5, report.FMT_PERCENT)
        )
        assert summary["Ending Headcount"] == (62, "#,##0")

    def test_summary_value_pairs_are_typed(self, model: dict) -> None:
        for key, (value, fmt) in model["summary"].items():
            assert isinstance(key, str)
            assert isinstance(fmt, str)
            assert isinstance(value, (int, float))

    def test_tax_is_zero_when_ebit_negative(self, model: dict) -> None:
        _, tax, _, _ = _row_by_label(model, "Tax")
        assert tax == [0.0] * 5

    def test_tax_applied_only_on_positive_ebit(self) -> None:
        a = _patched_assumptions(
            starting_revenue=1_000_000.0,
            revenue_growth=0.1,
            cogs_pct=0.3,
            starting_headcount=10,
            headcount_growth=0.0,
            cost_per_head=100_000.0,
            opex_pct={},
            depreciation=0.0,
        )
        with patch.object(report, "ASSUMPTIONS", a):
            model = report.compute_model()
        _, ebit, _, _ = _row_by_label(model, "EBIT")
        _, tax, _, _ = _row_by_label(model, "Tax")
        _, net, _, _ = _row_by_label(model, "Net Income")
        assert ebit == pytest.approx([-300_000.0, -230_000.0, -153_000.0, -68_300.0, 24_870.0])
        assert tax == pytest.approx([0.0, 0.0, 0.0, 0.0, 24_870.0 * 0.21])
        assert net == pytest.approx([e - t for e, t in zip(ebit, tax)])

    def test_zero_revenue_yields_zero_margins(self) -> None:
        a = _patched_assumptions(starting_revenue=0.0, years=1)
        with patch.object(report, "ASSUMPTIONS", a):
            model = report.compute_model()
        assert model["revenue"] == [0.0]
        assert _margin_by_label(model, "Gross Margin")[1] == [0.0]
        assert _margin_by_label(model, "EBITDA Margin")[1] == [0.0]
        assert _margin_by_label(model, "Net Margin")[1] == [0.0]

    def test_cagr_is_zero_for_single_year(self) -> None:
        a = _patched_assumptions(years=1)
        with patch.object(report, "ASSUMPTIONS", a):
            model = report.compute_model()
        assert model["summary"]["Revenue CAGR"] == pytest.approx((0.0, report.FMT_PERCENT))
        assert model["years"] == [2025]

    def test_summary_uses_model_wide_averages(self, model: dict) -> None:
        _, gross = _margin_by_label(model, "Gross Margin")
        _, net = _margin_by_label(model, "Net Margin")
        assert model["summary"]["Avg Gross Margin"] == pytest.approx(
            (sum(gross) / 5, report.FMT_PERCENT)
        )
        assert model["summary"]["Avg Net Margin"] == pytest.approx(
            (sum(net) / 5, report.FMT_PERCENT)
        )
        assert model["summary"]["Ending Headcount"] == (model["headcount"][-1], "#,##0")


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------


class TestStyleHelpers:
    def test_style_title_sets_value_font_fill_alignment(self, ws: Worksheet) -> None:
        cell = ws["A1"]
        report.style_title(cell, "Title")
        assert cell.value == "Title"
        assert cell.font.name == "Calibri"
        assert cell.font.size == 16
        assert cell.font.bold is True
        assert cell.font.color.rgb == "00FFFFFF"
        assert _fill(ws, "A1") == report.BRAND
        assert cell.alignment.horizontal == "left"
        assert cell.alignment.vertical == "center"
        assert cell.alignment.indent == 1

    def test_style_header_sets_brand_fill_and_border(self, ws: Worksheet) -> None:
        cell = ws["B3"]
        report.style_header(cell, "Header")
        assert cell.value == "Header"
        assert cell.font.bold is True
        assert cell.font.color.rgb == "00FFFFFF"
        assert _fill(ws, "B3") == report.BRAND
        assert cell.alignment.horizontal == "center"
        assert cell.alignment.vertical == "center"
        assert cell.border == report.BORDER_ALL

    def test_style_label_default(self, ws: Worksheet) -> None:
        cell = ws["A1"]
        report.style_label(cell, "Label")
        assert cell.value == "Label"
        assert cell.font.bold is False
        assert cell.alignment.horizontal == "left"
        assert cell.alignment.indent == 0
        assert cell.border == report.BORDER_ALL

    def test_style_label_bold_and_indented(self, ws: Worksheet) -> None:
        cell = ws["A1"]
        report.style_label(cell, "Sub", bold=True, indent=2)
        assert cell.font.bold is True
        assert cell.alignment.indent == 2

    def test_style_number_default_format(self, ws: Worksheet) -> None:
        cell = ws["B1"]
        report.style_number(cell, 1234.0)
        assert cell.value == 1234.0
        assert cell.number_format == report.FMT_CURRENCY
        assert cell.font.bold is False
        assert cell.alignment.horizontal == "right"
        assert cell.border == report.BORDER_ALL

    def test_style_number_custom_format_and_bold(self, ws: Worksheet) -> None:
        cell = ws["B1"]
        report.style_number(cell, 0.28, report.FMT_PERCENT, bold=True)
        assert cell.number_format == report.FMT_PERCENT
        assert cell.font.bold is True

    def test_set_widths(self, ws: Worksheet) -> None:
        report.set_widths(ws, {"A": 30.0, "B": 18.0, "C": 44.0})
        assert ws.column_dimensions["A"].width == 30
        assert ws.column_dimensions["B"].width == 18
        assert ws.column_dimensions["C"].width == 44

    def test_merge_title_merges_styles_anchor_and_sets_row_height(self, ws: Worksheet) -> None:
        report.merge_title(ws, "A1:C1", "Merged Title")
        assert _merged(ws) == {"A1:C1"}
        anchor = ws["A1"]
        assert anchor.value == "Merged Title"
        assert anchor.font.size == 16
        assert anchor.font.color.rgb == "00FFFFFF"
        assert _fill(ws, "A1") == report.BRAND
        assert ws.row_dimensions[1].height == 26


# ---------------------------------------------------------------------------
# build_assumptions — Sheet 1
# ---------------------------------------------------------------------------


class TestBuildAssumptions:
    def test_widths_and_title_band(self, ws: Worksheet) -> None:
        report.build_assumptions(ws)
        assert ws.column_dimensions["A"].width == 30
        assert ws.column_dimensions["B"].width == 18
        assert ws.column_dimensions["C"].width == 44
        assert _merged(ws) == {"A1:C1", "A15:C15"}
        assert ws["A1"].value == "Acme Robotics, Inc. — Model Assumptions"
        assert ws.row_dimensions[1].height == 26

    def test_header_row(self, ws: Worksheet) -> None:
        report.build_assumptions(ws)
        assert ws["A3"].value == "Parameter"
        assert ws["B3"].value == "Value"
        assert ws["C3"].value == "Notes"
        assert _fill(ws, "A3") == report.BRAND
        assert ws["A3"].border == report.BORDER_ALL

    def test_param_rows_exact_values_and_formats(self, ws: Worksheet) -> None:
        report.build_assumptions(ws)
        expected = [
            ("Start Year", 2025, "0", "First projected year"),
            ("Projection Horizon", 5, '0" yrs"', "Number of years modeled"),
            ("Starting Revenue", 4_800_000.0, report.FMT_CURRENCY, "Year-1 top line"),
            ("Revenue Growth (YoY)", 0.28, report.FMT_PERCENT, "Compounded annually"),
            ("COGS % of Revenue", 0.42, report.FMT_PERCENT, "Cost of goods sold"),
            ("Tax Rate", 0.21, report.FMT_PERCENT, "Applied to positive EBIT"),
            ("Starting Headcount", 32, "#,##0", "Year-1 employees"),
            ("Headcount Growth (YoY)", 0.18, report.FMT_PERCENT, "Hiring pace"),
            ("Cost per Head", 145_000.0, report.FMT_CURRENCY, "Fully-loaded annual"),
            ("Depreciation & Amort.", 240_000.0, report.FMT_CURRENCY, "Flat per year"),
        ]
        for i, (label, value, fmt, note) in enumerate(expected):
            row = 4 + i
            assert ws[f"A{row}"].value == label
            assert ws[f"B{row}"].value == value
            assert ws[f"B{row}"].number_format == fmt
            assert ws[f"C{row}"].value == note

    def test_param_notes_are_italic_grey(self, ws: Worksheet) -> None:
        report.build_assumptions(ws)
        for row in range(4, 14):
            cell = ws[f"C{row}"]
            assert cell.font.italic is True
            assert cell.font.color.rgb == "00808080"
            assert cell.alignment.horizontal == "left"
            assert cell.alignment.indent == 1
            assert cell.border == report.BORDER_ALL

    def test_zebra_banding_on_even_param_rows(self, ws: Worksheet) -> None:
        report.build_assumptions(ws)
        for row in range(4, 14):
            for col in ("A", "B", "C"):
                fill = _fill(ws, f"{col}{row}")
                if row % 2 == 0:
                    assert fill == report.ZEBRA
                else:
                    assert fill != report.ZEBRA

    def test_opex_subtable(self, ws: Worksheet) -> None:
        report.build_assumptions(ws)
        assert ws["A15"].value == "Operating Expense Ratios"
        assert _fill(ws, "A15") == report.BRAND
        assert ws["A16"].value == "Opex Line"
        assert ws["B16"].value == "% of Revenue"
        assert ws["C16"].value == "Notes"
        expected = [
            ("Sales & Marketing", 0.16),
            ("Research & Development", 0.11),
            ("General & Admin", 0.07),
        ]
        for i, (label, pct) in enumerate(expected):
            row = 17 + i
            assert ws[f"A{row}"].value == label
            assert ws[f"B{row}"].value == pct
            assert ws[f"B{row}"].number_format == report.FMT_PERCENT
            assert ws[f"C{row}"].value == "Scales with revenue"
            assert ws[f"C{row}"].font.italic is True

    def test_freeze_panes(self, ws: Worksheet) -> None:
        report.build_assumptions(ws)
        assert ws.freeze_panes == "A4"


# ---------------------------------------------------------------------------
# _style_pnl_row / _build_pnl_rows — P&L rendering internals
# ---------------------------------------------------------------------------


class TestStylePnlRow:
    def test_indented_label_is_stripped_and_indented(self, ws: Worksheet) -> None:
        row = ("  Sales & Marketing", [1.0, 2.0, 3.0, 4.0, 5.0], report.FMT_CURRENCY, False)
        report._style_pnl_row(ws, 7, 7, row)
        cell = ws.cell(row=7, column=1)
        assert cell.value == "Sales & Marketing"
        assert cell.alignment.indent == 1

    def test_plain_label_is_not_indented(self, ws: Worksheet) -> None:
        row = ("COGS", [1.0, 2.0, 3.0, 4.0, 5.0], report.FMT_CURRENCY, False)
        report._style_pnl_row(ws, 5, 7, row)
        assert ws.cell(row=5, column=1).alignment.indent == 0

    def test_totals_column_is_sum_of_years_and_bold(self, ws: Worksheet) -> None:
        row = ("Revenue", [1.0, 2.0, 3.0, 4.0, 5.0], report.FMT_CURRENCY, True)
        report._style_pnl_row(ws, 4, 7, row)
        total = ws.cell(row=4, column=7)
        assert total.value == 15.0
        assert total.number_format == report.FMT_CURRENCY
        assert total.font.bold is True

    def test_bold_row_gets_brand_light_fill_across_all_columns(self, ws: Worksheet) -> None:
        row = ("Revenue", [1.0, 2.0, 3.0, 4.0, 5.0], report.FMT_CURRENCY, True)
        report._style_pnl_row(ws, 4, 7, row)
        for col in range(1, 8):
            assert _rgb(ws.cell(row=4, column=col).fill) == report.BRAND_LIGHT

    def test_even_non_bold_row_gets_zebra_fill(self, ws: Worksheet) -> None:
        row = ("COGS", [1.0, 2.0, 3.0, 4.0, 5.0], report.FMT_CURRENCY, False)
        report._style_pnl_row(ws, 8, 7, row)
        for col in range(1, 8):
            assert _rgb(ws.cell(row=8, column=col).fill) == report.ZEBRA

    def test_odd_non_bold_row_gets_no_fill(self, ws: Worksheet) -> None:
        row = ("COGS", [1.0, 2.0, 3.0, 4.0, 5.0], report.FMT_CURRENCY, False)
        report._style_pnl_row(ws, 5, 7, row)
        for col in range(1, 8):
            assert ws.cell(row=5, column=col).fill.patternType is None


class TestBuildPnlRows:
    def test_returns_revenue_row_and_next_free_row(self, ws: Worksheet, model: dict) -> None:
        revenue_row, next_row = report._build_pnl_rows(ws, model, 7, 4)
        assert revenue_row == 4
        assert next_row == 4 + len(model["rows"])
        assert ws.cell(row=4, column=1).value == "Revenue"
        assert ws.cell(row=16, column=1).value == "Net Income"
        assert ws.cell(row=17, column=1).value is None


# ---------------------------------------------------------------------------
# _build_margin_table / build_model — Sheet 2 (the P&L)
# ---------------------------------------------------------------------------


class TestBuildMarginTable:
    def test_headers_years_avg_and_return_values(self, ws: Worksheet, model: dict) -> None:
        margin_hdr, margin_first, margin_last = report._build_margin_table(
            ws, model, model["years"], 7, 18
        )
        assert (margin_hdr, margin_first, margin_last) == (18, 19, 21)
        assert ws.cell(row=18, column=1).value == "Margins"
        assert ws.cell(row=18, column=1).border == report.BORDER_ALL
        for i, yr in enumerate(model["years"]):
            assert ws.cell(row=18, column=2 + i).value == str(yr)
        assert ws.cell(row=18, column=7).value == "Avg"

    def test_margin_values_are_percent_formatted_with_average(
        self, ws: Worksheet, model: dict
    ) -> None:
        report._build_margin_table(ws, model, model["years"], 7, 18)
        for r, (label, values) in enumerate(model["margin_rows"], start=19):
            assert ws.cell(row=r, column=1).value == label
            for i, v in enumerate(values):
                cell = ws.cell(row=r, column=2 + i)
                assert cell.value == v
                assert cell.number_format == report.FMT_PERCENT
            avg = ws.cell(row=r, column=7)
            assert avg.value == sum(values) / len(values)
            assert avg.number_format == report.FMT_PERCENT


class TestBuildModel:
    def test_returns_revenue_row_and_margin_header_row(self, ws: Worksheet, model: dict) -> None:
        assert report.build_model(ws, model) == (4, 18)

    def test_column_widths(self, ws: Worksheet, model: dict) -> None:
        report.build_model(ws, model)
        assert ws.column_dimensions["A"].width == 26
        for letter in "BCDEF":
            assert ws.column_dimensions[letter].width == 14
        assert ws.column_dimensions["G"].width == 16

    def test_title_band_and_header_row(self, ws: Worksheet, model: dict) -> None:
        report.build_model(ws, model)
        assert _merged(ws) == {"A1:G1"}
        assert ws["A1"].value == "5-Year P&L Projection (USD)"
        assert ws["A3"].value == "Line Item"
        assert [ws[f"{c}3"].value for c in "BCDEF"] == ["2025", "2026", "2027", "2028", "2029"]
        assert ws["G3"].value == "Total"

    def test_revenue_row_values_and_total(self, ws: Worksheet, model: dict) -> None:
        report.build_model(ws, model)
        for i, v in enumerate(model["revenue"]):
            cell = ws.cell(row=4, column=2 + i)
            assert cell.value == v
            assert cell.number_format == report.FMT_CURRENCY
            assert cell.font.bold is True
        total = ws["G4"]
        assert total.value == sum(model["revenue"])
        assert total.font.bold is True
        assert _fill(ws, "A4") == report.BRAND_LIGHT

    def test_net_income_is_last_pnl_row(self, ws: Worksheet, model: dict) -> None:
        report.build_model(ws, model)
        for i, v in enumerate(model["net_income"]):
            assert ws.cell(row=16, column=2 + i).value == v
        assert ws.cell(row=16, column=7).value == sum(model["net_income"])

    def test_formula_reference_note_row(self, ws: Worksheet, model: dict) -> None:
        report.build_model(ws, model)
        label = ws["A23"]
        assert label.value == "Formula ref (not live):"
        assert label.font.italic is True
        assert label.font.color.rgb == "00808080"
        ref = ws["B23"]
        assert ref.value == "Gross Margin = Gross Profit / Revenue  (row 4)"
        assert ref.font.italic is True

    def test_margin_subtable_position(self, ws: Worksheet, model: dict) -> None:
        report.build_model(ws, model)
        assert ws["A18"].value == "Margins"
        assert ws["A19"].value == "Gross Margin"
        assert ws["A21"].value == "Net Margin"
        assert ws["A22"].value is None

    def test_conditional_formatting_color_scale_on_margins(
        self, ws: Worksheet, model: dict
    ) -> None:
        report.build_model(ws, model)
        rule = _rule(ws, "B19:F21")
        assert rule.type == "colorScale"
        cfvo = rule.colorScale.cfvo
        assert [f.type for f in cfvo] == ["min", "percentile", "max"]
        assert cfvo[1].val == 50
        assert [c.rgb for c in rule.colorScale.color] == ["00F8696B", "00FFEB84", "0063BE7B"]

    def test_conditional_formatting_data_bar_on_revenue_row(
        self, ws: Worksheet, model: dict
    ) -> None:
        report.build_model(ws, model)
        rule = _rule(ws, "B4:F4")
        assert rule.type == "dataBar"
        assert rule.dataBar.color.rgb == "00" + report.BRAND
        assert rule.dataBar.showValue is True
        assert [f.type for f in rule.dataBar.cfvo] == ["min", "max"]

    def test_freeze_panes_locks_header_row_and_label_column(
        self, ws: Worksheet, model: dict
    ) -> None:
        report.build_model(ws, model)
        assert ws.freeze_panes == "B4"


# ---------------------------------------------------------------------------
# build_dashboard — Sheet 3
# ---------------------------------------------------------------------------


class TestBuildDashboard:
    @pytest.fixture
    def sheets(self) -> tuple[Worksheet, Worksheet]:
        wb = Workbook()
        return wb.active, wb.create_sheet("Model")

    def test_widths_and_title_band(self, sheets: tuple[Worksheet, Worksheet]) -> None:
        ws, model_ws = sheets
        report.build_dashboard(ws, model_ws, report.compute_model(), 4, 18)
        assert ws.column_dimensions["A"].width == 26
        assert ws.column_dimensions["B"].width == 18
        assert _merged(ws) == {"A1:E1"}
        assert ws["A1"].value == "Executive Dashboard"
        assert ws.row_dimensions[1].height == 26

    def test_kpi_header(self, sheets: tuple[Worksheet, Worksheet]) -> None:
        ws, model_ws = sheets
        report.build_dashboard(ws, model_ws, report.compute_model(), 4, 18)
        assert ws["A3"].value == "Key Metric"
        assert ws["B3"].value == "Value"
        assert _fill(ws, "A3") == report.BRAND

    def test_kpi_rows_exact_values_and_formats(self, sheets: tuple[Worksheet, Worksheet]) -> None:
        ws, model_ws = sheets
        model = report.compute_model()
        report.build_dashboard(ws, model_ws, model, 4, 18)
        expected = [
            ("Total Revenue (5y)", 41_759_551.488, report.FMT_CURRENCY),
            ("Total Net Income (5y)", -24_527_707.642_88, report.FMT_CURRENCY),
            ("Revenue CAGR", 0.28, report.FMT_PERCENT),
            ("Avg Gross Margin", 0.58, report.FMT_PERCENT),
            ("Avg Net Margin", -0.623_274_718_523_025_5, report.FMT_PERCENT),
            ("Ending Headcount", 62, "#,##0"),
        ]
        for i, (label, value, fmt) in enumerate(expected):
            row = 4 + i
            assert ws[f"A{row}"].value == label
            assert ws[f"B{row}"].value == pytest.approx(value)
            assert ws[f"B{row}"].number_format == fmt
            assert ws[f"B{row}"].font.bold is True

    def test_zebra_banding_on_even_kpi_rows(self, sheets: tuple[Worksheet, Worksheet]) -> None:
        ws, model_ws = sheets
        report.build_dashboard(ws, model_ws, report.compute_model(), 4, 18)
        for row in range(4, 10):
            for col in ("A", "B"):
                cell = ws[f"{col}{row}"]
                if row % 2 == 0:
                    assert _rgb(cell.fill) == report.ZEBRA
                else:
                    assert cell.fill.patternType is None

    def test_data_bars_on_kpi_values(self, sheets: tuple[Worksheet, Worksheet]) -> None:
        ws, model_ws = sheets
        report.build_dashboard(ws, model_ws, report.compute_model(), 4, 18)
        rule = _rule(ws, "B4:B9")
        assert rule.type == "dataBar"
        assert rule.dataBar.color.rgb == "00" + report.ACCENT

    def test_bar_chart_revenue_vs_net_income(self, sheets: tuple[Worksheet, Worksheet]) -> None:
        ws, model_ws = sheets
        model = report.compute_model()
        report.build_dashboard(ws, model_ws, model, 4, 18)
        bar = ws._charts[0]
        assert isinstance(bar, report.BarChart)
        assert bar.type == "col"
        assert _title_text(bar.title) == "Revenue vs Net Income"
        assert _title_text(bar.y_axis.title) == "USD"
        assert _title_text(bar.x_axis.title) == "Year"
        assert bar.height == 8
        assert bar.width == 18
        assert bar.anchor == "D3"
        assert len(bar.series) == 2
        assert _series_range(bar, 0) == "'Model'!$B$4:$F$4"
        assert _series_range(bar, 1) == "'Model'!$B$16:$F$16"
        assert bar.series[0].title.strRef.f == "'Model'!A4"
        assert bar.series[1].title.strRef.f == "'Model'!A16"
        assert bar.series[0].cat.numRef.f == "'Model'!$B$3:$F$3"

    def test_bar_chart_net_income_row_is_last_pnl_row(
        self, sheets: tuple[Worksheet, Worksheet]
    ) -> None:
        ws, model_ws = sheets
        model = report.compute_model()
        n_rows = len(model["rows"])
        revenue_row = 4
        report.build_dashboard(ws, model_ws, model, revenue_row, 18)
        assert (
            _series_range(ws._charts[0], 1)
            == f"'Model'!$B${revenue_row + n_rows - 1}:$F${revenue_row + n_rows - 1}"
        )

    def test_line_chart_net_margin_trend(self, sheets: tuple[Worksheet, Worksheet]) -> None:
        ws, model_ws = sheets
        report.build_dashboard(ws, model_ws, report.compute_model(), 4, 18)
        line = ws._charts[1]
        assert isinstance(line, report.LineChart)
        assert _title_text(line.title) == "Net Margin Trend"
        assert _title_text(line.y_axis.title) == "Margin"
        assert _title_text(line.x_axis.title) == "Year"
        assert line.anchor == "D19"
        assert len(line.series) == 1
        assert _series_range(line, 0) == "'Model'!$B$21:$F$21"
        assert line.series[0].title.strRef.f == "'Model'!A21"
        assert line.series[0].smooth is True

    def test_line_chart_net_margin_row_is_margin_hdr_plus_3(
        self, sheets: tuple[Worksheet, Worksheet]
    ) -> None:
        ws, model_ws = sheets
        model = report.compute_model()
        margin_hdr = 18
        report.build_dashboard(ws, model_ws, model, 4, margin_hdr)
        assert _series_range(ws._charts[1], 0) == f"'Model'!$B${margin_hdr + 3}:$F${margin_hdr + 3}"

    def test_freeze_panes(self, sheets: tuple[Worksheet, Worksheet]) -> None:
        ws, model_ws = sheets
        report.build_dashboard(ws, model_ws, report.compute_model(), 4, 18)
        assert ws.freeze_panes == "A4"


# ---------------------------------------------------------------------------
# main — workbook assembly + save
# ---------------------------------------------------------------------------


class TestMain:
    def test_main_saves_three_sheet_workbook_active_on_dashboard(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        out = tmp_path / "report.xlsx"
        monkeypatch.setattr(sys, "argv", ["report.py", str(out)])
        report.main()
        assert out.exists()
        wb = load_workbook(out)
        assert wb.sheetnames == ["Assumptions", "Model", "Dashboard"]
        assert wb.active.title == "Dashboard"
        assert wb["Assumptions"]["B6"].value == 4_800_000.0
        assert wb["Model"]["G4"].value == pytest.approx(41_759_551.488)
        assert wb["Dashboard"]["B9"].value == 62

    def test_main_defaults_to_out_xlsx_in_cwd(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        monkeypatch.chdir(tmp_path)
        monkeypatch.setattr(sys, "argv", ["report.py"])
        report.main()
        assert (tmp_path / "out.xlsx").exists()
