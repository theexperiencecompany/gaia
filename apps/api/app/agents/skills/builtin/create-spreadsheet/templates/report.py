"""FINANCIAL MODEL TEMPLATE (openpyxl) — a 5-year P&L projection + dashboard.

This is benchmark-quality few-shot reference material. It demonstrates the full
openpyxl surface an LLM needs to build a polished workbook:

  * Styling      — Font / PatternFill / Alignment / Border / Side
  * Numbers      — number_format (currency '"$"#,##0', percent '0.0%')
  * Layout       — column_dimensions[..].width, freeze_panes, merged title cells
  * Highlighting — conditional formatting (ColorScaleRule + DataBarRule)
  * Charts       — openpyxl.chart BarChart + LineChart with Reference/categories

CRITICAL CONTRACT — NO LIVE FORMULAS.
The sandbox has no spreadsheet recalc engine, so a written `=SUM(...)` shows up
blank when opened. Every authoritative number below is computed in *Python* and
the resulting value is written into the cell. Where a formula is instructive we
write it as a plain string into a clearly-labeled, separate "formula" cell so a
reader can see the math — but it is never the source of truth.

Run via: bash scripts/build.sh report.py out.xlsx
"""

import sys

from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# --- Brand palette & reusable number-format codes -----------------------------
# Constants over magic values: every color / format string lives here once.
BRAND = "1A4D8F"  # deep blue — headers
BRAND_LIGHT = "D6E4F5"  # pale blue — banor / total emphasis
ACCENT = "2E8B57"  # green — positive callouts
ZEBRA = "F2F6FB"  # subtle row banding
WHITE = "FFFFFF"
GREY = "808080"

FMT_CURRENCY = '"$"#,##0'  # $1,234  (no decimals — clean for a P&L)
FMT_CURRENCY_K = '"$"#,##0,"k"'  # $1,234k (scaled thousands)
FMT_PERCENT = "0.0%"  # 12.3%
FMT_MULTIPLE = '0.00"x"'  # 1.45x  (e.g. for ratios)

THIN = Side(style="thin", color="C8D2E0")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# --- Assumptions (the only inputs a user edits) -------------------------------
# A single dict drives the whole model. Change these numbers and rerun.
ASSUMPTIONS = {
    "company": "Acme Robotics, Inc.",
    "start_year": 2025,
    "years": 5,
    "starting_revenue": 4_800_000.0,  # Year-1 revenue
    "revenue_growth": 0.28,  # YoY growth rate
    "cogs_pct": 0.42,  # COGS as % of revenue
    "tax_rate": 0.21,  # corporate tax on positive EBIT
    "starting_headcount": 32,
    "headcount_growth": 0.18,  # YoY headcount growth
    "cost_per_head": 145_000.0,  # fully-loaded annual cost / employee
    # Opex line items as % of revenue (salaries handled separately via headcount)
    "opex_pct": {
        "Sales & Marketing": 0.16,
        "Research & Development": 0.11,
        "General & Admin": 0.07,
    },
    "depreciation": 240_000.0,  # flat annual D&A
}


# --- DRY styling helpers ------------------------------------------------------
def style_title(cell: Cell, text: str) -> None:
    """Large brand-colored title used for merged header blocks."""
    cell.value = text
    cell.font = Font(name="Calibri", size=16, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=BRAND)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)


def style_header(cell: Cell, text: str) -> None:
    """Column / section header — white bold text on the brand fill."""
    cell.value = text
    cell.font = Font(bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=BRAND)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER_ALL


def style_label(cell: Cell, text: str, *, bold: bool = False, indent: int = 0) -> None:
    """Left-aligned row label in the first column."""
    cell.value = text
    cell.font = Font(bold=bold)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=indent)
    cell.border = BORDER_ALL


def style_number(cell: Cell, value: float, fmt: str = FMT_CURRENCY, *, bold: bool = False) -> None:
    """Right-aligned numeric cell with a number format. Value is pre-computed."""
    cell.value = value
    cell.number_format = fmt
    cell.font = Font(bold=bold)
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.border = BORDER_ALL


def set_widths(ws: Worksheet, widths: dict[str, float]) -> None:
    """Apply explicit column widths via column_dimensions."""
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def merge_title(ws: Worksheet, span: str, text: str) -> None:
    """Merge a cell range into one title band and style its anchor cell."""
    ws.merge_cells(span)
    anchor = span.split(":", maxsplit=1)[0]
    style_title(ws[anchor], text)
    ws.row_dimensions[ws[anchor].row].height = 26


# --- Core computation (pure Python — produces every authoritative number) -----
def compute_model() -> dict:
    """Build the 5-year P&L. Returns a dict of label -> list-of-yearly-values
    plus the year labels and derived summary metrics. No spreadsheet math."""
    a = ASSUMPTIONS
    n = a["years"]
    years = [a["start_year"] + i for i in range(n)]

    # Revenue compounds off the starting figure at the growth rate.
    revenue = [a["starting_revenue"] * (1 + a["revenue_growth"]) ** i for i in range(n)]
    cogs = [r * a["cogs_pct"] for r in revenue]
    gross_profit = [r - c for r, c in zip(revenue, cogs)]

    # Opex: each %-of-revenue line, plus a headcount-driven personnel line.
    headcount = [
        round(a["starting_headcount"] * (1 + a["headcount_growth"]) ** i) for i in range(n)
    ]
    personnel = [hc * a["cost_per_head"] for hc in headcount]
    opex_lines: dict[str, list[float]] = {"Personnel": personnel}
    for name, pct in a["opex_pct"].items():
        opex_lines[name] = [r * pct for r in revenue]

    total_opex = [sum(opex_lines[k][i] for k in opex_lines) for i in range(n)]
    ebitda = [gp - ox for gp, ox in zip(gross_profit, total_opex)]
    depreciation = [a["depreciation"]] * n
    ebit = [e - d for e, d in zip(ebitda, depreciation)]
    # Tax only on positive EBIT (a simple but realistic guard).
    tax = [max(0.0, e) * a["tax_rate"] for e in ebit]
    net_income = [e - t for e, t in zip(ebit, tax)]

    # Derived margins (ratios computed in Python, written as percent values).
    gross_margin = [gp / r if r else 0.0 for gp, r in zip(gross_profit, revenue)]
    ebitda_margin = [e / r if r else 0.0 for e, r in zip(ebitda, revenue)]
    net_margin = [ni / r if r else 0.0 for ni, r in zip(net_income, revenue)]

    # Compound annual growth rate of revenue across the projected periods.
    rev_cagr = (revenue[-1] / revenue[0]) ** (1 / (n - 1)) - 1 if n > 1 else 0.0

    return {
        "years": years,
        "rows": [
            ("Revenue", revenue, FMT_CURRENCY, True),
            ("COGS", cogs, FMT_CURRENCY, False),
            ("Gross Profit", gross_profit, FMT_CURRENCY, True),
            # Opex breakdown (indented under a section header handled at render).
            *[(f"  {k}", v, FMT_CURRENCY, False) for k, v in opex_lines.items()],
            ("Total Opex", total_opex, FMT_CURRENCY, True),
            ("EBITDA", ebitda, FMT_CURRENCY, True),
            ("Depreciation & Amort.", depreciation, FMT_CURRENCY, False),
            ("EBIT", ebit, FMT_CURRENCY, True),
            ("Tax", tax, FMT_CURRENCY, False),
            ("Net Income", net_income, FMT_CURRENCY, True),
        ],
        "margin_rows": [
            ("Gross Margin", gross_margin),
            ("EBITDA Margin", ebitda_margin),
            ("Net Margin", net_margin),
        ],
        "summary": {
            "Total Revenue (5y)": (sum(revenue), FMT_CURRENCY),
            "Total Net Income (5y)": (sum(net_income), FMT_CURRENCY),
            "Revenue CAGR": (rev_cagr, FMT_PERCENT),
            "Avg Gross Margin": (sum(gross_margin) / n, FMT_PERCENT),
            "Avg Net Margin": (sum(net_margin) / n, FMT_PERCENT),
            "Ending Headcount": (headcount[-1], "#,##0"),
        },
        "revenue": revenue,
        "net_income": net_income,
        "net_margin": net_margin,
        "headcount": headcount,
    }


# --- Sheet 1: Assumptions -----------------------------------------------------
def build_assumptions(ws: Worksheet) -> None:
    a = ASSUMPTIONS
    set_widths(ws, {"A": 30, "B": 18, "C": 44})
    merge_title(ws, "A1:C1", f"{a['company']} — Model Assumptions")

    style_header(ws["A3"], "Parameter")
    style_header(ws["B3"], "Value")
    style_header(ws["C3"], "Notes")

    # (label, value, number_format, note)
    params = [
        ("Start Year", a["start_year"], "0", "First projected year"),
        ("Projection Horizon", a["years"], '0" yrs"', "Number of years modeled"),
        ("Starting Revenue", a["starting_revenue"], FMT_CURRENCY, "Year-1 top line"),
        ("Revenue Growth (YoY)", a["revenue_growth"], FMT_PERCENT, "Compounded annually"),
        ("COGS % of Revenue", a["cogs_pct"], FMT_PERCENT, "Cost of goods sold"),
        ("Tax Rate", a["tax_rate"], FMT_PERCENT, "Applied to positive EBIT"),
        ("Starting Headcount", a["starting_headcount"], "#,##0", "Year-1 employees"),
        ("Headcount Growth (YoY)", a["headcount_growth"], FMT_PERCENT, "Hiring pace"),
        ("Cost per Head", a["cost_per_head"], FMT_CURRENCY, "Fully-loaded annual"),
        ("Depreciation & Amort.", a["depreciation"], FMT_CURRENCY, "Flat per year"),
    ]
    row = 4
    for label, value, fmt, note in params:
        style_label(ws[f"A{row}"], label)
        style_number(ws[f"B{row}"], value, fmt)
        ws[f"C{row}"] = note
        ws[f"C{row}"].font = Font(italic=True, color=GREY)
        ws[f"C{row}"].alignment = Alignment(horizontal="left", indent=1)
        ws[f"C{row}"].border = BORDER_ALL
        # Zebra banding on even rows for readability.
        if row % 2 == 0:
            for col in ("A", "B", "C"):
                ws[f"{col}{row}"].fill = PatternFill("solid", fgColor=ZEBRA)
        row += 1

    # Opex %-of-revenue sub-table.
    row += 1
    merge_title(ws, f"A{row}:C{row}", "Operating Expense Ratios")
    row += 1
    style_header(ws[f"A{row}"], "Opex Line")
    style_header(ws[f"B{row}"], "% of Revenue")
    style_header(ws[f"C{row}"], "Notes")
    row += 1
    for name, pct in a["opex_pct"].items():
        style_label(ws[f"A{row}"], name)
        style_number(ws[f"B{row}"], pct, FMT_PERCENT)
        ws[f"C{row}"] = "Scales with revenue"
        ws[f"C{row}"].font = Font(italic=True, color=GREY)
        ws[f"C{row}"].border = BORDER_ALL
        row += 1

    ws.freeze_panes = "A4"  # keep title + header visible while scrolling


# --- Sheet 2: Model (the P&L) -------------------------------------------------
def _style_pnl_row(ws: Worksheet, r: int, total_col: int, row: tuple) -> None:
    """Render a single P&L line item with subtotal/zebra emphasis."""
    label, values, fmt, bold = row
    is_indented = label.startswith("  ")
    style_label(
        ws.cell(row=r, column=1),
        label.strip() if is_indented else label,
        bold=bold,
        indent=1 if is_indented else 0,
    )
    for i, v in enumerate(values):
        style_number(ws.cell(row=r, column=2 + i), v, fmt, bold=bold)
    # Bold totals column = sum across years (computed in Python).
    style_number(ws.cell(row=r, column=total_col), sum(values), fmt, bold=True)
    # Emphasize subtotal rows with a light fill.
    if bold:
        for c in range(1, total_col + 1):
            cell = ws.cell(row=r, column=c)
            if cell.fill.fgColor.rgb in (None, "00000000"):
                cell.fill = PatternFill("solid", fgColor=BRAND_LIGHT)
    elif r % 2 == 0:  # zebra banding on non-subtotal rows
        for c in range(1, total_col + 1):
            ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=ZEBRA)


def _build_pnl_rows(ws: Worksheet, model: dict, total_col: int, data_start: int) -> tuple[int, int]:
    """Render all P&L line items. Returns (revenue_row, next_free_row)."""
    revenue_row = 0
    r = data_start
    for row in model["rows"]:
        if row[0] == "Revenue":
            revenue_row = r
        _style_pnl_row(ws, r, total_col, row)
        r += 1
    return revenue_row, r


def _build_margin_table(
    ws: Worksheet, model: dict, years: list, total_col: int, start_row: int
) -> tuple[int, int, int]:
    """Render the margin sub-table. Returns (margin_hdr, margin_first, margin_last)."""
    margin_hdr = start_row
    style_header(ws.cell(row=start_row, column=1), "Margins")
    for i, yr in enumerate(years):
        style_header(ws.cell(row=start_row, column=2 + i), str(yr))
    style_header(ws.cell(row=start_row, column=total_col), "Avg")
    r = start_row + 1
    margin_first = r
    for label, values in model["margin_rows"]:
        style_label(ws.cell(row=r, column=1), label)
        for i, v in enumerate(values):
            style_number(ws.cell(row=r, column=2 + i), v, FMT_PERCENT)
        style_number(ws.cell(row=r, column=total_col), sum(values) / len(values), FMT_PERCENT)
        r += 1
    return margin_hdr, margin_first, r - 1


def build_model(ws: Worksheet, model: dict) -> tuple[int, int]:
    """Render the P&L. Returns the 1-based row index of the Revenue row so the
    dashboard charts can build References into it."""
    years = model["years"]
    n = len(years)
    # Columns: A=label, B..=years, last=Total.
    total_col = 2 + n  # column index of the 'Total' column
    last_letter = get_column_letter(total_col)

    set_widths(ws, {"A": 26})
    for i in range(n):
        ws.column_dimensions[get_column_letter(2 + i)].width = 14
    ws.column_dimensions[last_letter].width = 16

    merge_title(ws, f"A1:{last_letter}1", "5-Year P&L Projection (USD)")

    # Header row (row 3): label blank, then each year, then 'Total'.
    hdr = 3
    style_header(ws.cell(row=hdr, column=1), "Line Item")
    for i, yr in enumerate(years):
        style_header(ws.cell(row=hdr, column=2 + i), str(yr))
    style_header(ws.cell(row=hdr, column=total_col), "Total")

    data_start = hdr + 1
    revenue_row, next_row = _build_pnl_rows(ws, model, total_col, data_start)

    # Margin sub-table below the P&L (percent-formatted, computed values).
    margin_hdr, margin_first, margin_last = _build_margin_table(
        ws, model, years, total_col, next_row + 1
    )
    r = margin_last + 1

    # An instructive (non-authoritative) formula string — clearly labeled.
    note_row = r + 1
    ws.cell(row=note_row, column=1, value="Formula ref (not live):").font = Font(
        italic=True, color=GREY
    )
    ws.cell(
        row=note_row,
        column=2,
        value=f"Gross Margin = Gross Profit / Revenue  (row {revenue_row})",
    ).font = Font(italic=True, color=GREY)

    # Conditional formatting: 3-color scale across the margin block so a reader
    # instantly sees which years are strongest. Applied to the year columns.
    cf_range = f"{get_column_letter(2)}{margin_first}:{get_column_letter(1 + n)}{margin_last}"
    ws.conditional_formatting.add(
        cf_range,
        ColorScaleRule(
            start_type="min",
            start_color="F8696B",  # red (low)
            mid_type="percentile",
            mid_value=50,
            mid_color="FFEB84",  # yellow
            end_type="max",
            end_color="63BE7B",  # green (high)
        ),
    )

    # Data bars on the Revenue row across the year columns — visual magnitude.
    rev_bar_range = f"{get_column_letter(2)}{revenue_row}:{get_column_letter(1 + n)}{revenue_row}"
    ws.conditional_formatting.add(
        rev_bar_range,
        DataBarRule(start_type="min", end_type="max", color=BRAND, showValue=True),
    )

    # Freeze the header row AND the label column.
    ws.freeze_panes = ws.cell(row=data_start, column=2).coordinate

    return revenue_row, margin_hdr  # margin_hdr used by dashboard line chart


# --- Sheet 3: Summary / Dashboard --------------------------------------------
def build_dashboard(
    ws: Worksheet, model_ws: Worksheet, model: dict, revenue_row: int, margin_hdr: int
) -> None:
    years = model["years"]
    n = len(years)
    set_widths(ws, {"A": 26, "B": 18})
    merge_title(ws, "A1:E1", "Executive Dashboard")

    # KPI cards (label / value), each value pre-computed in Python.
    style_header(ws["A3"], "Key Metric")
    style_header(ws["B3"], "Value")
    r = 4
    metric_first = r
    for label, (value, fmt) in model["summary"].items():
        style_label(ws[f"A{r}"], label)
        style_number(ws[f"B{r}"], value, fmt, bold=True)
        if r % 2 == 0:
            ws[f"A{r}"].fill = PatternFill("solid", fgColor=ZEBRA)
            ws[f"B{r}"].fill = PatternFill("solid", fgColor=ZEBRA)
        r += 1
    metric_last = r - 1

    # Conditional formatting on the KPI value column — data bars for at-a-glance
    # comparison of the magnitude of each metric.
    ws.conditional_formatting.add(
        f"B{metric_first}:B{metric_last}",
        DataBarRule(start_type="min", end_type="max", color=ACCENT),
    )

    # --- Chart 1: BarChart — Revenue vs Net Income by year --------------------
    # References point at the MODEL sheet. Revenue is at `revenue_row`; Net
    # Income is the last row of the P&L block. We locate Net Income by scanning
    # the rows definition (it is always the final entry).
    net_income_row = (
        revenue_row + sum(1 for _ in model["rows"]) - 1
    )  # rows are contiguous; Net Income is last
    bar = BarChart()
    bar.type = "col"
    bar.title = "Revenue vs Net Income"
    bar.y_axis.title = "USD"
    bar.x_axis.title = "Year"
    bar.height = 8
    bar.width = 18
    # Two data series: Revenue row and Net Income row (include the label col so
    # titles_from_data picks up the series name from column A).
    rev_ref = Reference(
        model_ws, min_col=1, min_row=revenue_row, max_col=1 + n, max_row=revenue_row
    )
    ni_ref = Reference(
        model_ws,
        min_col=1,
        min_row=net_income_row,
        max_col=1 + n,
        max_row=net_income_row,
    )
    bar.add_data(rev_ref, titles_from_data=True, from_rows=True)
    bar.add_data(ni_ref, titles_from_data=True, from_rows=True)
    cats = Reference(model_ws, min_col=2, max_col=1 + n, min_row=3, max_row=3)
    bar.set_categories(cats)
    ws.add_chart(bar, "D3")

    # --- Chart 2: LineChart — Net Margin trend --------------------------------
    # Net Margin is the 3rd margin row (margin_hdr + 3).
    net_margin_row = margin_hdr + 3
    line = LineChart()
    line.title = "Net Margin Trend"
    line.y_axis.title = "Margin"
    line.x_axis.title = "Year"
    line.height = 8
    line.width = 18
    nm_ref = Reference(
        model_ws,
        min_col=1,
        min_row=net_margin_row,
        max_col=1 + n,
        max_row=net_margin_row,
    )
    line.add_data(nm_ref, titles_from_data=True, from_rows=True)
    line.set_categories(cats)
    series = list(line.series)
    if series:
        series[0].smooth = True
    ws.add_chart(line, "D19")

    ws.freeze_panes = "A4"


# --- Assemble the workbook ----------------------------------------------------
def main() -> None:
    model = compute_model()

    wb = Workbook()
    ws_assumptions = wb.active
    ws_assumptions.title = "Assumptions"
    ws_model = wb.create_sheet("Model")
    ws_dashboard = wb.create_sheet("Dashboard")

    build_assumptions(ws_assumptions)
    revenue_row, margin_hdr = build_model(ws_model, model)
    build_dashboard(ws_dashboard, ws_model, model, revenue_row, margin_hdr)

    # Open on the dashboard for an executive-friendly first impression.
    wb.active = wb.sheetnames.index("Dashboard")

    out = sys.argv[1] if len(sys.argv) > 1 else "out.xlsx"
    wb.save(out)


if __name__ == "__main__":
    main()
